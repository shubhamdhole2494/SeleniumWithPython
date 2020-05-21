import openpyxl
import openpyxl.worksheet.filters
import openpyxl.cell
from openpyxl import cell

file = "/Users/sdhole/Documents/python_workspace/SeleniumWithPython/Excel_operation/excelUtils/Test.xlsx"
# Open Sheet
wb = openpyxl.load_workbook(file)

print(wb.sheetnames)

# Create New Sheet
# wb.create_sheet()
# wb.save(file)

ws = wb.active

# Rename Sheet
# for i in wb.sheetnames:
#     if i == 'Sheet':
#         i = wb['Sheet1']
#     else:
#         print('Sheet is not found')
#
#     wb.save(file)

print(ws.max_column)
print(ws.max_row)


# Get column Data
# for item in range(1,ws.max_row+1):
#     #print(ws.cell(column=1,row=item).value)
#     if ws.cell(column=1,row=item).value == 'Shubham':
#         cl = ws.cell.Cell.coordinate()
#         print(cl)
#     else:
#         print("none")
def get_coordinate():
    for item in range(1, ws.max_row + 1):
        # print(ws.cell(column=1,row=item).value)
        if ws.cell(column=1, row=item).value == 'Shubham':
            return ws.cell(column=1, row=item).row


co_or = get_coordinate()
print(co_or)

for item in range(1,ws.max_row+1):
    print(ws.cell(row=co_or,column=item).value)

# ws.auto_filter.ref = "A1:D4"
# ws.auto_filter.add_filter_column(0,['Shubham'])
#
# wb.save(file)
# print(ws.max_column)
# print(ws.max_row)
